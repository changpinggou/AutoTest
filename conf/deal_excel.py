# _*_ coding:UTF-8 _*_
"""
@project -> File :DigitalHumanTest -> com_func 
@Author: Dora
@Date: 2023/9/3 15:28
@Desc:
1-功能描述: (1)update_excel()更新parans.xlsx;(2)get_data()获取测试数据
2-实现步骤：
3-状态（废弃/使用）：
"""
import os, sys, re
from openpyxl import load_workbook

PROJ_ROOT = os.path.dirname(os.path.abspath(__file__))
PROJ_PARENT_ROOT = os.path.abspath(os.path.dirname(PROJ_ROOT))
parentdir = os.path.dirname(PROJ_ROOT)
sys.path.insert(0, parentdir)

# 将控制台输出存入日志文件
from conf.pytest_log import Logger
Logger = Logger()
logger = Logger.loggering('test_digital_human')

class DealExcel:
    def __init__(self,sheet_name):
        self.excel_path = os.path.join(PROJ_PARENT_ROOT, "conf", "test_data.xlsx")
        print(f'excel_path: {self.excel_path}')
        self.sheet_name = sheet_name
        # 获取工作薄
        self.workbook = load_workbook(self.excel_path)
        # 选择工作表
        self.sheet = self.workbook[self.sheet_name]

    def set_data_format(self, cell_value):
        # 调整读到的单元格数据为能被脚本使用的格式
        # cell_value格式: self.sheet[cell_name].value
        return_data = ''
        if not cell_value:
            return_data = None
        else:
            pattern = r"\(.*?\)"  # 使用正则表达式非贪婪模式匹配括号及其内部的任意内容
            matches = re.findall(pattern, cell_value)
            list_data = []
            if matches:
                for match in matches:
                    tuple_result = tuple(match.strip('(').strip(')').split(','))
                    # print(f'tuple_result: {tuple_result}')
                    # data in params.xlsx: (a,b,c),(e,f,g)
                    # return data: [('a','b','c'),('e','f','g')],input data to script by group
                    list_data.append(tuple_result)
                    # print("The contents in parentheses:", match)
                    return_data = list_data
            elif "," in str(cell_value):
                # data in params.xlsx: a,b,c
                # return data: ['a','b','c']
                return_data = cell_value.split(',')
            else:
                return_data = cell_value
                #return_data.append(cell_value)
        #print(f'return_data: {return_data}, {type(return_data)}')
        return return_data

    def get_right_cells(self, cell, n):
        # 获取指定单元格右边第n格的位置
        column = cell.column
        row = cell.row
        right_column = column + n
        right_cell = self.sheet.cell(row=row, column=right_column)
        #print(f'right_cell: {right_cell}')
        return right_cell

    def get_position(self, cell_data):
        # 获取数据在表中的单元格位置
        # 读取整个工作表的数据到矩阵
        matrix = list(self.sheet.iter_rows(values_only=True))
        # 获取case_name所在单元格
        # 遍历矩阵中的每一行
        cell_position = ''
        for row_index, row in enumerate(matrix, start=1):
            # 遍历当前行的每一列
            for column_index, value in enumerate(row, start=1):
                # 判断值是否等于目标值
                if value == cell_data:
                    # 获取单元格的位置
                    cell_position = self.sheet.cell(row=row_index, column=column_index).coordinate
                    #print(f'cell_position: {cell_position}')
                    break
        return cell_position

    def get_data(self, case_name, params_name='',cell_num=1):
        # 获取case_name所在单元格
        cell_position = self.get_position(case_name)
        if not cell_position:
            print(f"the {case_name} is not in the sheet")
        else:
            print('cell position')
            cell = self.sheet[cell_position]
            #print(f'cell_value: {cell.value}')
            # 获取用例测试数据
            if params_name == 'quality':
                return self.set_data_format(self.get_right_cells(cell, 1).value)
            elif params_name == 'video_path_to_model':
                return self.set_data_format(self.get_right_cells(cell, 2).value)
            elif params_name == 'video_to_model':
                return self.set_data_format(self.get_right_cells(cell, 3).value)
            elif params_name == 'result_model':
                return self.set_data_format(self.get_right_cells(cell, 4).value)
            elif params_name == 'video_path_to_inference':
                return self.set_data_format(self.get_right_cells(cell, 5).value)
            elif params_name == 'video_to_inference':
                return self.set_data_format(self.get_right_cells(cell, 6).value)
            elif params_name == 'label_config_base64':
                return self.set_data_format(self.get_right_cells(cell, 7).value)
            elif params_name == 'result_inference':
                return self.set_data_format(self.get_right_cells(cell, 8).value)
            elif params_name == 'input_model':
                return self.set_data_format(self.get_right_cells(cell, 9).value)
            elif params_name == 'input_inference':
                return self.set_data_format(self.get_right_cells(cell, 10).value)
            elif params_name == 'audio_path':
                return self.set_data_format(self.get_right_cells(cell, 11).value)
            elif params_name == 'audio_name':
                return self.set_data_format(self.get_right_cells(cell, 12).value)
            elif params_name == 'result_video':
                return self.set_data_format(self.get_right_cells(cell, 13).value)
            elif params_name == 'test_result':
                print('this_params_name: ' + 'test_result')
                print(str(self.get_right_cells(cell, 14).value))
                return self.set_data_format(self.get_right_cells(cell, 14).value)
            # 获取传入的变量的值，digital_server，output那些
            elif params_name in ['digital_server','output']:
                return self.set_data_format(self.get_right_cells(cell, 1).value)
            else:
                logger.warning(f"the {params_name} is not in the sheet")

        # 保存Excel文件
        self.workbook.save(self.excel_path)
        # 关闭Excel文件
        self.workbook.close()

    def update_excel(self, case_name, data,params_name=''):
        # 获取case_name所在单元格
        cell_position = self.get_position(case_name)
        if not cell_position:
            logger.warning("warning: the 'case_name' is not in the sheet")
            print("warning: the 'case_name' is not in the sheet")
        else:
            cell = self.sheet[cell_position]
            # 获取 cell_position单元格的右边数据
            if params_name == 'result_model':
                """
                # add to data,simultaneously keep the old data
                # get the old data
                current_time = datetime.now().strftime("%m-%d %H:%M:%S")
                exist_data = self.get_right_cells(self.sheet, cell, 6).value
                if exist_data:
                    data = data + '<--'+current_time +'|'+ exist_data
                """
                # update cell value to new data directly
                self.get_right_cells(cell, 4).value = data
            elif params_name == 'result_inference':
                # update cell value to new data directly
                self.get_right_cells(cell, 8).value = data
            elif params_name == 'result_video':
                # update cell value to new data directly
                self.get_right_cells(cell, 13).value = data
            elif params_name == 'test_result':
                # update cell value to new data directly
                self.get_right_cells(cell, 14).value = data
            else:
                self.get_right_cells(cell,1).value = data
        # 保存Excel文件
        self.workbook.save(self.excel_path)
        # 关闭Excel文件
        self.workbook.close()


if __name__ == '__main__':

    digital_server = 'digital_server-v1.6.0.142-202309080839-feature-69cf6c0f-2023-09-08-21-47'
    output = '6path/to/success'
    case_scope = '6SMOKE_CASES'
    excel = DealExcel(sheet_name = 'params')
    print(excel.sheet_name)
    excel.update_excel(case_name='digital_server',data=digital_server)
    excel.update_excel(case_name='output',data=output)

    excel2 = DealExcel(sheet_name = 'API_CASES')
    print(excel.sheet_name)
    #excel2.update_excel(case_name='digital_server',data=digital_server)
    excel2.update_excel(case_name='output',data=output)
    value = excel2.get_data(case_name='test_great_change_serial_create',params_name='video_to_model')
    print(f'type(value): {type(value)}')
    print(f'value: {value}')
    print('###########')
    excel3 = DealExcel(sheet_name='ALL_CASES')
    case_name = 'test_create_model'
    quality = excel3.get_data(case_name=case_name,params_name='quality')
    print(f'quality: {quality}')
    video_path_to_model = excel3.get_data(case_name=case_name,params_name='video_path_to_model')
    print(f'video_path_to_model: {video_path_to_model}')
    video_to_model = excel3.get_data(case_name=case_name,params_name='video_to_model')
    print(f'video_to_model: {video_to_model}')
    print('###################')
    result_model = excel3.get_data(case_name=case_name,params_name='result_model')
    print(f'type(result_model): {type(result_model)}')
    print(f'result_model: {result_model}')


